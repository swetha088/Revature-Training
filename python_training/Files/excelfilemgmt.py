from openpyxl import Workbook, load_workbook
import os

def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Age"])
    sheet.append(["John Doe", 30])
    sheet.append(["Jane Smith", 25])
    workbook.save(filename)

def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(f"Name: {row[0]}, Age: {row[1]}")

# Function to delete an Excel file
def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f"{filename} deleted successfully")
    else:
        print(f"{filename} does not exist")

def delete_excel2(filename):
    try:
        os.remove(filename)
        print(f"{filename} deleted successfully")
    except FileNotFoundError:
        print(f"{filename} does not exist")
    except PermissionError:
        print(f"Permission error: {filename} is currently in use")


filename = "data.xlsx"
write_excel(filename)
print("Data read from Excel file:")
read_excel(filename)
delete_excel2(filename)








